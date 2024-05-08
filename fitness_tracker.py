import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
import csv
import numpy as np
import os
import pandas as pd
from tkinter import simpledialog
import time

csv_directory = os.path.expanduser("~/Documents/files")
fitness_data_file = os.path.expanduser("~/Documents/fitness_data.xlsx")


def add_column_headings_if_not_exist(file_path):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
        if not all(ws.cell(row=1, column=i).value in ["Name", "Age", "Weight (kg)", "Height (m)", "Gender",
                                                      "Food Category", "Physical Issues", "BMI", "BMI Category",
                                                      ]
                for i in range(1, 10)):
            raise Exception("Column headings missing")
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Weight (kg)", "Height (m)", "Gender", "Food Category", "Physical Issues", "BMI",
                   "BMI Category"])
        wb.save(filename=file_path)


def add_registration_details_to_excel(registration_details):
    try:
        documents_folder = os.path.expanduser("~/Documents")
        file_path = os.path.join(documents_folder, "fitness_data.xlsx")
        add_column_headings_if_not_exist(file_path)

        wb = load_workbook(filename=file_path)
        ws = wb.active

        ws.append(registration_details)  # Append new registration details

        # Calculate BMI and BMI category
        weight, height = registration_details[2], registration_details[3]
        bmi = weight / (height ** 2)
        bmi_category = calculate_bmi_category(bmi)

        # Append BMI and BMI category to the last column
        ws.cell(row=ws.max_row, column=8).value = bmi
        ws.cell(row=ws.max_row, column=9).value = bmi_category

        wb.save(filename=file_path)
        messagebox.showinfo("Registration", "User registered successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"Error occurred while saving data: {e}")


def calculate_bmi_category(bmi):
    if bmi < 18.5:
        return "underweight"
    elif 18.5 <= bmi < 25:
        return "normal"
    elif 25 <= bmi < 30:
        return "overweight"
    else:
        return "obese"


def get_user_data(name, ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and row[0].lower() == name.lower():
            return row
    return None


def set_day_goals():
    try:
        name = input("Enter your name: ")
        documents_folder = os.path.expanduser("~/Documents")
        file_path = os.path.join(documents_folder, "fitness_data.xlsx")
        wb = load_workbook(filename=file_path)
        ws = wb.active
        user_data = get_user_data(name, ws)
        if user_data:
            day_number = int(input("Enter the day number for setting goals (1-10): "))
            if 1 <= day_number <= 10:
                food_category = str(user_data[5]) if user_data[5] else "Unknown"
                bmi_category = str(user_data[8]).lower() if user_data[8] else "Unknown"

                csv_filename = f"{food_category}_{bmi_category}.csv"
                csv_file_path = os.path.join(csv_directory, csv_filename)

                try:
                    with open(csv_file_path, "r") as csv_file:
                        lines = csv_file.readlines()
                    if 1 <= day_number <= len(lines):
                        headings = lines[0].strip().split(",")
                        data = lines[day_number].strip().split(",")

                        table_window = tk.Toplevel()
                        table_window.title(f"Day {day_number} Goals for {name}")
                        table_window.geometry("1200x200")
                        center_window(table_window)

                        tree = ttk.Treeview(table_window, columns=headings, show="headings")
                        tree.pack(expand=True, fill="both")

                        for col in headings:
                            tree.heading(col, text=col)

                        tree.insert("", "end", values=data)
                    else:
                        messagebox.showerror("Error", "Invalid day number.")
                except FileNotFoundError:
                    messagebox.showerror("Error", f"File '{csv_filename}' not found in directory: {csv_directory}")
            else:
                messagebox.showerror("Error", "Invalid day number. Please enter a number between 1 and 10.")
        else:
            messagebox.showerror("Error", "User not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Error occurred: {e}")



def track_progress():
    try:
        name = input("Enter your name: ")
        documents_folder = os.path.expanduser("~/Documents")
        file_path = os.path.join(documents_folder, "fitness_data.xlsx")
        wb = load_workbook(filename=file_path)
        ws = wb.active
        user_data = get_user_data(name, ws)
        if user_data:
            day_number = int(input("Enter the day number for progress tracking (1-10): "))
            if 1 <= day_number <= 10:
                # Ensure food_category and bmi_category are not None before using them
                food_category = user_data[5] if user_data[5] else "Unknown"
                bmi_category = user_data[8].lower() if user_data[8] else "Unknown"  # Adjusted index to 8

                csv_filename = f"{food_category}_{bmi_category}.csv"
                csv_file_path = os.path.join(csv_directory, csv_filename)

                # Check if the CSV file exists before trying to open it
                if os.path.exists(csv_file_path):
                    with open(csv_file_path, "r") as csv_file:
                        csv_reader = csv.reader(csv_file)
                        next(csv_reader)  # Skip the header row
                        for _ in range(day_number):
                            row = next(csv_reader)
                        water_target = float(row[9])  # Index 9 corresponds to the water target column
                        calorie_target = float(row[10])  # Index 10 corresponds to the calorie target column

                    water_cups = float(input("Enter your water intake for the day (in cups): "))
                    diet_attendance_percentage = float(input("Enter the percentage of diet attended (0-100): "))
                    exercise_hours = float(input("Enter the number of hours of exercise: "))

                    water_intake = water_cups * 236.588
                    calorie_intake = calorie_target * (diet_attendance_percentage / 100)
                    exercise_minutes = exercise_hours * 60  # Convert hours to minutes
                    exercise_progress = (exercise_minutes / (3 * 60)) * 100  # Assuming 3 hours of exercise is the goal

                    # Plot progress
                    plot_progress(water_target, water_intake, calorie_target, calorie_intake, exercise_progress, 100)

                else:
                    messagebox.showerror("Error",
                                         f"CSV file '{csv_filename}' not found in directory: {csv_directory}")
            else:
                messagebox.showerror("Error", "Invalid day number. Please enter a number between 1 and 10.")
        else:
            messagebox.showerror("Error", "User not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Error occurred: {e}")


def plot_progress(water_target, water_progress, calorie_target, calorie_progress, exercise_target, exercise_progress):
    categories = ['Water', 'Calorie', 'Exercise']
    targets = [water_target, calorie_target, exercise_target]
    progresses = [water_progress, calorie_progress, exercise_progress]

    num_categories = len(categories)
    index = np.arange(num_categories)

    bar_width = 0.20

    fig, ax = plt.subplots()

    bars_targets = ax.bar(index, targets, bar_width, label='Target', color='blue')
    bars_progress = ax.bar(index + bar_width, progresses, bar_width, label='Progress', color='orange')

    ax.set_xlabel('Categories')
    ax.set_ylabel('Values')
    ax.set_title('Progress Tracking')
    ax.set_xticks(index + bar_width / 2)
    ax.set_xticklabels(categories)
    ax.legend()

    # Calculate average percentage of progress
    average_percentage = (water_progress / water_target + calorie_progress / calorie_target + exercise_progress) / 3

    # Set motivational message based on average percentage
    if average_percentage >= 90:
        motivational_message = "Great job! Keep up the excellent work!"
    elif average_percentage >= 70:
        motivational_message = "You're doing well! Keep pushing yourself!"
    else:
        motivational_message = "Stay focused and keep working towards your goals!"

    # Add motivational message below the chart
    ax.text(0.5, -0.15, motivational_message, horizontalalignment='center', verticalalignment='center',
            transform=ax.transAxes, fontsize=10)

    plt.show()


def register_user():
    def add_registration():
        try:
            name = name_entry.get()
            age = int(age_entry.get())
            weight = float(weight_entry.get())
            height = float(height_entry.get())
            gender = gender_var.get()
            food_category = food_category_var.get()
            physical_issues = physical_issues_entry.get()

            registration_data = [name, age, weight, height, gender, food_category, physical_issues]
            add_registration_details_to_excel(registration_data)
            messagebox.showinfo("Registration", "User registered successfully!")
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")

    registration_window = tk.Toplevel()
    registration_window.title("Register User")
    registration_window.geometry("400x500")
    center_window(registration_window)

    name_label = tk.Label(registration_window, text="Name:")
    name_label.pack()
    name_entry = tk.Entry(registration_window)
    name_entry.pack()

    age_label = tk.Label(registration_window, text="Age:")
    age_label.pack()
    age_entry = tk.Entry(registration_window)
    age_entry.pack()

    weight_label = tk.Label(registration_window, text="Weight (kg):")
    weight_label.pack()
    weight_entry = tk.Entry(registration_window)
    weight_entry.pack()

    height_label = tk.Label(registration_window, text="Height (m):")
    height_label.pack()
    height_entry = tk.Entry(registration_window)
    height_entry.pack()

    gender_label = tk.Label(registration_window, text="Gender:")
    gender_label.pack()
    gender_var = tk.StringVar(value="Male")
    gender_radio1 = tk.Radiobutton(registration_window, text="Male", variable=gender_var, value="Male")
    gender_radio1.pack()
    gender_radio2 = tk.Radiobutton(registration_window, text="Female", variable=gender_var, value="Female")
    gender_radio2.pack()

    food_category_label = tk.Label(registration_window, text="Food Category:")
    food_category_label.pack()
    food_category_var = tk.StringVar(value="veg")
    food_category_radio1 = tk.Radiobutton(registration_window, text="veg", variable=food_category_var, value="veg")
    food_category_radio1.pack()
    food_category_radio2 = tk.Radiobutton(registration_window, text="nonveg", variable=food_category_var,
                                          value="nonveg")
    food_category_radio2.pack()

    physical_issues_label = tk.Label(registration_window, text="Physical Issues:")
    physical_issues_label.pack()
    physical_issues_entry = tk.Entry(registration_window)
    physical_issues_entry.pack()

    register_button = tk.Button(registration_window, text="Register", command=add_registration)
    register_button.pack()

    registration_window.mainloop()


def know_your_bmi_category():
    name = input("Enter your name: ")
    documents_folder = os.path.expanduser("~/Documents")
    file_path = os.path.join(documents_folder, "fitness_data.xlsx")
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
        user_data = get_user_data(name, ws)
        if user_data:
            bmi = user_data[7]  # Index 7 corresponds to the BMI column (column H)
            bmi_category = user_data[8]  # Index 8 corresponds to the BMI Category column (column I)
            print(f"Your BMI is: {bmi}")
            print(f"Your BMI Category is: {bmi_category}")
        else:
            print("User not found.")
            response = input("Do you want to enter your height and weight to calculate BMI? (yes/no): ").lower()
            if response == "yes":
                height = float(input("Enter your height (in meters): "))
                weight = float(input("Enter your weight (in kg): "))
                bmi = weight / (height ** 2)
                bmi_category = calculate_bmi_category(bmi)
                print(f"Your BMI is: {bmi}")
                print(f"Your BMI Category is: {bmi_category}")
            elif response == "no":
                print("Exiting BMI calculator.")
            else:
                print("Invalid response. Exiting BMI calculator.")
    except ValueError:
        print("Invalid input. Please enter a valid number.")
    except Exception as e:
        print(f"Error occurred: {e}")


def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f"{width}x{height}+{x}+{y}")


'''def share_your_progress():
    try:
        name = input("Enter your name: ")
        documents_folder = os.path.expanduser("~/Documents")
        file_path = os.path.join(documents_folder, "fitness_data.xlsx")
        wb = load_workbook(filename=file_path)
        ws = wb.active
        user_data = get_user_data(name, ws)
        if user_data:
            day_number = int(input("Enter the day number for progress tracking (1-10): "))
            if 1 <= day_number <= 10:
                water_cups = float(input("Enter your water intake for the day (in cups): "))
                diet_attendance_percentage = float(input("Enter the percentage of diet attended (0-100): "))
                exercise_hours = float(input("Enter the number of hours of exercise: "))
                save_progress(name, day_number, water_cups, diet_attendance_percentage, exercise_hours)
                messagebox.showinfo("Success", "Progress shared successfully!")
            else:
                messagebox.showerror("Error", "Invalid day number. Please enter a number between 1 and 10.")
        else:
            messagebox.showerror("Error", "User not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Error occurred: {e}")


def save_progress(name, day_number, water_cups, diet_attendance_percentage, exercise_hours):
    try:
        documents_folder = os.path.expanduser("~/Documents")
        file_path = os.path.join(documents_folder, "progress.csv")
        with open(file_path, "a") as file:
            writer = csv.writer(file)
            writer.writerow([name, day_number, water_cups, diet_attendance_percentage, exercise_hours])
    except Exception as e:
        messagebox.showerror("Error", f"Error occurred while saving progress: {e}")'''

import tkinter as tk
from tkinter import messagebox
import time

def show_reminder(message):
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showinfo("Reminder", message)

def set_reminder():
    # Function to set reminders for water intake and exercise
    
    # Ask the user for reminder timings
    water_timing_1 = simpledialog.askstring("Water Reminder", "Enter timing for water reminder 1 (HH:MM AM/PM):")
    water_timing_2 = simpledialog.askstring("Water Reminder", "Enter timing for water reminder 2 (HH:MM AM/PM):")
    exercise_timing_1 = simpledialog.askstring("Exercise Reminder", "Enter timing for exercise reminder 1 (HH:MM AM/PM):")
    exercise_timing_2 = simpledialog.askstring("Exercise Reminder", "Enter timing for exercise reminder 2 (HH:MM AM/PM):")
    
    # Convert timings to 24-hour format
    water_timing_1_24 = time.strptime(water_timing_1, "%I:%M %p")
    water_timing_2_24 = time.strptime(water_timing_2, "%I:%M %p")
    exercise_timing_1_24 = time.strptime(exercise_timing_1, "%I:%M %p")
    exercise_timing_2_24 = time.strptime(exercise_timing_2, "%I:%M %p")
    
    # Set up reminders using a loop
    while True:
        current_time = time.localtime()
        
        # Check if it's time for the water reminder 1
        if current_time.tm_hour == water_timing_1_24.tm_hour and current_time.tm_min == water_timing_1_24.tm_min:
            show_reminder("Reminder: Drink water!")
        
        # Check if it's time for the water reminder 2
        if current_time.tm_hour == water_timing_2_24.tm_hour and current_time.tm_min == water_timing_2_24.tm_min:
            show_reminder("Reminder: Drink water!")
        
        # Check if it's time for the exercise reminder 1
        if current_time.tm_hour == exercise_timing_1_24.tm_hour and current_time.tm_min == exercise_timing_1_24.tm_min:
            show_reminder("Reminder: Exercise!")
        
        # Check if it's time for the exercise reminder 2
        if current_time.tm_hour == exercise_timing_2_24.tm_hour and current_time.tm_min == exercise_timing_2_24.tm_min:
            show_reminder("Reminder: Exercise!")
        
        # Delay for 1 minute before checking again
        time.sleep(60)








def main():
    root = tk.Tk()
    root.title("Fitness Tracker")

    register_button = tk.Button(root, text="NEW User? Register", command=register_user)
    register_button.pack()

    bmi_button = tk.Button(root, text="Know Your BMI Category", command=know_your_bmi_category)
    bmi_button.pack()

    set_goals_button = tk.Button(root, text="Set Day Goals", command=set_day_goals)
    set_goals_button.pack()

    track_button = tk.Button(root, text="Track Progress", command=track_progress)
    track_button.pack()

    reminder_button = tk.Button(root, text="Set Reminder", command=set_reminder)
    reminder_button.pack()

    '''share_button = tk.Button(root, text="Share Your Progress", command=share_your_progress)
    share_button.pack()'''

    exit_button = tk.Button(root, text="Exit", command=root.destroy)
    exit_button.pack()

   


    root.mainloop()


if __name__ == "__main__":
    main()
